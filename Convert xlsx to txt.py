import pandas as pd
from openpyxl import load_workbook
import datetime

writeTo = open("Tier List Ratings " + str(datetime.date.today()) + ".txt", "w")
wb = load_workbook("Tier List.xlsx")

colours = ["Red", "Blue", "Green", "Yellow"]

for colourNum in range(len(colours)):

     print("Writing " + colours[colourNum] + ", please wait...")
     
     writeTo.write("#" + colours[colourNum] + " Units\n\n")
     writeTo.write("Name | Rating | Obtain Method | 6* in GL?\n")
     writeTo.write(":--|:--|:--|:--\n")

     sheet = wb.get_sheet_by_name(colours[colourNum])
     for x in range(sheet.max_row):
          writeTo.write(sheet['A' + str(x+1)].value.replace('https', 'http') + " | " + str(sheet['B' + str(x+1)].value) + " | " + sheet['C' + str(x+1)].value + " | " + sheet['D' + str(x+1)].value + "\n")
     writeTo.write("\n")

writeTo.close()

input("Writing complete!")
