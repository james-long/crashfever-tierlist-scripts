from lxml import html
import requests
from openpyxl import load_workbook

#star = tree.xpath('//table[2]//td[2]/text()')
#image = tree.xpath('//table[2]//td[1]/a/img/@src')

wb = load_workbook("Tier List.xlsx")

sheetNames = ["Red", "Blue", "Green", "Yellow"]
sheetURLs = ["https://game8.jp/crashfever/40065", "https://game8.jp/crashfever/40266", "https://game8.jp/crashfever/40259", "https://game8.jp/crashfever/40060"]

for colourNum in range(len(sheetNames)):

     print("Comparing " + sheetNames[colourNum] + ", please wait...")
     
     sheet = wb.get_sheet_by_name(sheetNames[colourNum])
     page = requests.get(sheetURLs[colourNum])
     tree = html.fromstring(page.content)
     gameLinks = tree.xpath('//table[2]//td[1]/a/@href')
     gameLinkNums = [i.split('/')[-1] for i in gameLinks]
     gameNames = tree.xpath('//table[2]//td[1]/a/text()')
     gameRatings = tree.xpath('//table[2]//td[3]/b/text()')
     for x in range(sheet.max_row):
          fullName = sheet['A' + str(x+1)].value
          if(fullName.find("https://") >= 0):
               currRating = str(sheet['B' + str(x+1)].value)[:2]
               linkToMatch = fullName[fullName.find("https://"):fullName.rfind(")")]
               linkToMatchNum = linkToMatch.split('/')[-1]
               for num, link in enumerate(gameLinkNums):
                    if(linkToMatchNum == link):
                         gameRating = gameRatings[num][:2]
                         if(gameRating.isdigit()):
                              if(gameRating != currRating):
                                   print("Rating of " + gameRating + " (game8) does not match rating of " + currRating + " (tierlist) - " + fullName[:fullName.find("https://")-1])

input("Analysis complete!")
               

